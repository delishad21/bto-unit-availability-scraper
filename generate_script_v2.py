import os
import gzip
import json
import time
import openpyxl
import xml.etree.ElementTree as ET
from datetime import datetime, date
from collections import defaultdict
from datetime import datetime
from seleniumwire import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenPyxlImage
from PIL import Image as PILImage
from webdriver_manager.chrome import ChromeDriverManager

# === CONFIGURATION ===
BTO_URL = "" # Replace with your BTO project URL
# Example: BTO_URL = "https://homes.hdb.gov.sg/home/bto/details/2024-06_BTO_JSHFsjhfsjFSJHFsk"
BTO_UNIT_PREFIX = "" # Replace with your BTO unit prefix. You can find this in the URL of the BTO project page.
# Example: if your URL is "https://homes.hdb.gov.sg/home/bto/details/2024-06_BTO_JSHFsjhfsjFSJHFsk", then BTO_UNIT_PREFIX = "2024-06_BTO_"

# API and XML endpoints
TARGET_JSON_URL = "https://homes.hdb.gov.sg/home-api/protected/v1/newFlat/getSelectionProjectAvailabilityAndEthnic"
TARGET_XML_URL = "https://homes.hdb.gov.sg/home-api/protected/v1/newFlat/getProtectedUnitXml"
DATA_HISTORY_PATH = "availability_log.json"
IMAGE_PATH = "layout.png"

def login_and_capture_requests():
    options = Options()
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

    api_data, xml_unit_ids, project_name, room_type = None, set(), None, None

    try:
        driver.get("https://iam.hdb.gov.sg/common/login")
        print("Scan QR with SingPass to login.")
        WebDriverWait(driver, 60).until(EC.url_contains("https://services2.hdb.gov.sg/webapp"))
        print("Login successful.")

        driver.get(BTO_URL)
        time.sleep(1)
        driver.get(BTO_URL)
        time.sleep(2)

        for request in driver.requests:
            if request.response and request.url.startswith(TARGET_JSON_URL) and request.method == "POST":
                body = request.response.body
                if 'gzip' in request.response.headers.get('Content-Encoding', ''):
                    body = gzip.decompress(body)
                api_data = json.loads(body.decode('utf-8'))
                break

        for request in driver.requests:
            if request.response and request.url.startswith(TARGET_XML_URL) and request.method == "POST":
                body = request.response.body
                if 'gzip' in request.response.headers.get('Content-Encoding', ''):
                    body = gzip.decompress(body)
                root = ET.fromstring(body.decode('utf-8'))
                project_name = root.find(".//project-name").text
                room_type = root.find(".//type").text
                for unit_elem in root.findall(".//unit-id"):
                    xml_unit_ids.add(unit_elem.text)
                break
    finally:
        driver.quit()

    return api_data, xml_unit_ids, project_name, room_type


def load_history():
    if os.path.exists(DATA_HISTORY_PATH):
        with open(DATA_HISTORY_PATH, 'r') as f:
            return json.load(f)
    return {"runs": []}


def update_history(history, availability_set, xml_unit_ids):
    timestamp = datetime.now().isoformat()
    today = date.today()

    # Find the most recent previous run that was not today
    previous_run = None
    for run in reversed(history.get("runs", [])):
        try:
            run_date = datetime.fromisoformat(run["timestamp"]).date()
            if run_date < today:
                previous_run = run
                break
        except Exception:
            continue

    if previous_run:
        raw_prev = previous_run.get("timestamp", "")
        try:
            dt_prev = datetime.fromisoformat(raw_prev)
            previous_timestamp_str = dt_prev.strftime("%d %B %Y, %H:%M:%S")
        except Exception:
            previous_timestamp_str = raw_prev
        last_run_avail = set(previous_run.get("available", []))
    else:
        previous_timestamp_str = ""
        last_run_avail = set()

    newly_taken = sorted(last_run_avail - availability_set)

    # Save the new run to history
    history["runs"].append({
        "timestamp": timestamp,
        "available": sorted(availability_set),
        "newly_taken": newly_taken,
        "previous_timestamp": previous_timestamp_str
    })

    with open(DATA_HISTORY_PATH, 'w') as f:
        json.dump(history, f, indent=2)

    # remove units that are not in the XML data
    newly_taken = [unit for unit in newly_taken if unit in xml_unit_ids]

    return newly_taken, previous_timestamp_str


def prepare_block_data(xml_unit_ids):
    block_units = defaultdict(lambda: defaultdict(set))
    all_floors = defaultdict(set)
    for unit_id in xml_unit_ids:
        parts = unit_id.split("_")
        if len(parts) != 5:
            continue
        _, _, block, floor, unit = parts
        block_units[block][unit].add(int(floor))
        all_floors[block].add(int(floor))
    return block_units, all_floors


def create_excel(api_data, xml_unit_ids, newly_taken, previous_timestamp_str, project_name, room_type):

    EXCEL_PATH = f"Unit_Availability_{project_name}_{room_type}_{datetime.today().strftime('%Y-%m-%d')}.xlsx"

    availability_set = set(api_data.get("availabilitySet", []))
    block_units, all_floors = prepare_block_data(xml_unit_ids)
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    white_fill = PatternFill("solid", fgColor="FFFFFF")
    grey_fill = PatternFill("solid", fgColor="AAAAAA")
    red_fill = PatternFill("solid", fgColor="FF0000")
    black_fill = PatternFill("solid", fgColor="000000")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"))
    full_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    thick_border = Border(left=Side(style="thick"), right=Side(style="thick"), top=Side(style="thick"), bottom=Side(style="thick"))
    bold_font = Font(bold=True)

    img_to_add = None
    if os.path.exists(IMAGE_PATH):
        with PILImage.open(IMAGE_PATH) as pil_img:
            orig_width, orig_height = pil_img.size
        scale_factor = min(1.0, (35 * 20) / orig_height)
        new_width = int(orig_width * scale_factor)
        new_height = int(orig_height * scale_factor)
        img_to_add = OpenPyxlImage(IMAGE_PATH)
        img_to_add.width = new_width
        img_to_add.height = new_height

    block_stats = {}

    for block, units in block_units.items():
        create_block_sheet(wb, block, units, all_floors[block], availability_set, newly_taken,
                           img_to_add, white_fill, grey_fill, red_fill, black_fill,
                           thin_border, full_border, thick_border, bold_font, block_stats)

    create_summary_sheet(wb, block_stats, newly_taken, previous_timestamp_str, img_to_add)

    wb._sheets = [wb[s] for s in ["_Summary"] + sorted([s for s in wb.sheetnames if s != "_Summary"], key=lambda x: x.lower())]

    os.makedirs("Generated_Files", exist_ok=True)
    wb.save(os.path.join("Generated_Files", EXCEL_PATH))
    print(f"✅ Workbook saved: {EXCEL_PATH}")
    os.startfile(os.path.join("Generated_Files", EXCEL_PATH))


def create_block_sheet(wb, block, units, floor_set, availability_set, newly_taken,
                       img_to_add, white_fill, grey_fill, red_fill, black_fill,
                       thin_border, full_border, thick_border, bold_font, block_stats):
    ws = wb.create_sheet(title=block)
    max_floor = max(floor_set) if floor_set else 1
    floors = sorted(set(range(1, max_floor + 1)), reverse=True)
    unit_labels = sorted(units.keys())

    for col_idx, unit in enumerate(unit_labels):
        label_col = col_idx * 2 + 2
        flag_col = label_col + 1
        cell = ws.cell(row=2, column=label_col, value=unit)
        cell.font = Font(bold=True, size=13)
        cell.border = thick_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(flag_col)].hidden = True

    ws.row_dimensions[2].height = 25
    ws.column_dimensions[get_column_letter(1)].width = 10
    ws.column_dimensions[get_column_letter(len(unit_labels) * 2 + 2)].width = 3

    taken = available = total = 0
    unit_taken_count = {unit: 0 for unit in unit_labels}
    unit_total_count = {unit: 0 for unit in unit_labels}

    for row_idx, floor in enumerate(floors, start=3):
        for col_idx, unit in enumerate(unit_labels):
            label_col = col_idx * 2 + 2
            flag_col = label_col + 1
            unit_id = f"{BTO_UNIT_PREFIX}{block}_{floor:02}_{unit}"

            cell = ws.cell(row=row_idx, column=label_col)
            flag_cell = ws.cell(row=row_idx, column=flag_col)

            if floor in units[unit]:
                cell.value = floor
                total += 1
                unit_total_count[unit] += 1
                if unit_id in availability_set:
                    cell.fill = white_fill
                    flag_cell.value = 0
                    available += 1
                elif unit_id in newly_taken:
                    cell.fill = red_fill
                    flag_cell.value = 1
                    taken += 1
                    unit_taken_count[unit] += 1
                else:
                    cell.fill = grey_fill
                    flag_cell.value = 1
                    taken += 1
                    unit_taken_count[unit] += 1
            else:
                cell.fill = black_fill
                flag_cell.value = -1

            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        ws.row_dimensions[row_idx].height = 20

    # Add per unit statistics
    stat_row = len(floors) + 3
    ws.cell(row=stat_row, column=1, value="Taken").font = bold_font
    ws.cell(row=stat_row, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=stat_row + 1, column=1, value="Left").font = bold_font
    ws.cell(row=stat_row + 1, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=stat_row + 2, column=1, value="Total").font = bold_font
    ws.cell(row=stat_row + 2, column=1).alignment = Alignment(horizontal="center")
    for col_idx, unit in enumerate(unit_labels):
        label_col = col_idx * 2 + 2
        ws.cell(row=stat_row, column=label_col, value=unit_taken_count[unit]).font = Font(bold=True)
        ws.cell(row=stat_row, column=label_col).alignment = Alignment(horizontal="center")
        ws.cell(row=stat_row + 1, column=label_col, value=unit_total_count[unit]- unit_taken_count[unit]).font = Font(bold=True)
        ws.cell(row=stat_row + 1, column=label_col).alignment = Alignment(horizontal="center")
        ws.cell(row=stat_row + 2, column=label_col, value=unit_total_count[unit]).font = Font(bold=True)
        ws.cell(row=stat_row + 2, column=label_col).alignment = Alignment(horizontal="center")


    # Add Block statistics
    left_buffer_col = 1
    right_buffer_col = len(unit_labels) * 2 + 2
    stat_col = right_buffer_col + 2
    add_block_stats(ws, stat_col, taken, available, total, bold_font, full_border)

    # Colour black buffer rows for top
    for c in range(1, right_buffer_col + 1):
        ws.cell(row=1, column=c).fill = black_fill

    for row in range(1, stat_row):
        ws.cell(row=row, column=left_buffer_col).fill = black_fill
        ws.cell(row=row, column=right_buffer_col).fill = black_fill

    if img_to_add:
        anchor_cell = f"{get_column_letter(stat_col)}8"
        img_copy = OpenPyxlImage(IMAGE_PATH)
        img_copy.width = img_to_add.width
        img_copy.height = img_to_add.height
        img_copy.anchor = anchor_cell
        ws.add_image(img_copy)

    block_stats[block] = {"total": total, "taken": taken, "available": available}


def add_block_stats(ws, stat_col, taken, available, total, font, border):

    ws.column_dimensions[get_column_letter(stat_col)].width = 12

    labels = ["Block Total", "Taken", "Left", "% Taken"]
    values = [total, taken, available, taken / total if total else 0]

    for i, (label, value) in enumerate(zip(labels, values), start=0):
        ws.cell(row=2 + i, column=stat_col, value=label).font = font
        ws.cell(row=2 + i, column=stat_col, value=label).border = border
        cell = ws.cell(row=2 + i, column=stat_col + 1, value=value)
        cell.font = font
        cell.border = border
        if i == 3:
            cell.number_format = '0.0%'


def create_summary_sheet(wb, block_stats, newly_taken, previous_timestamp_str, img):
    ws = wb.create_sheet(title="_Summary")
    headers = ["Block", "Total", "Taken", "Available", "% Taken", "% Available"]
    ws.append(headers)

    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True, size=13)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    row = 2
    total_all = taken_all = available_all = 0

    for block, stats in block_stats.items():
        total_all += stats["total"]
        taken_all += stats["taken"]
        available_all += stats["available"]

        ws.append([
            block,
            stats["total"],
            stats["taken"],
            stats["available"],
            stats["taken"] / stats["total"] if stats["total"] else 0,
            stats["available"] / stats["total"] if stats["total"] else 0
        ])

        for col in range(1, 7):
            # Increase col width for better visibility
            ws.column_dimensions[get_column_letter(col)].width = 18
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                                 top=Side(style="thin"), bottom=Side(style="thin"))
            if col in [5, 6]:
                cell.number_format = '0.0%'
        row += 1


    # Add totals row
    ws.append(["TOTAL", total_all, taken_all, available_all,
               taken_all / total_all if total_all else 0,
               available_all / total_all if total_all else 0])

    for col in range(1, 7):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=Side(style="thin"), bottom=Side(style="thin"))
        if col in [5, 6]:
            cell.number_format = '0.0%'

    # increase row height up till current row
    for r in range(1, row + 1):
        ws.row_dimensions[r].height = 20

    if img:
        summary_image = OpenPyxlImage(IMAGE_PATH)
        summary_image.width = img.width
        summary_image.height = img.height
        summary_image.anchor = "H2"
        ws.add_image(summary_image)

    row += 2
    ws.cell(row=row, column=1, value=f"Units Taken Since {previous_timestamp_str}").font = Font(bold=True, size=13)
    ws.append(["Block", "Level", "Unit"])
    for unit_str in newly_taken:
        parts = unit_str.split("_")
        if len(parts) == 5:
            ws.append([parts[2], parts[3], parts[4]])

# === Main Execution ===
if __name__ == "__main__":
    print("🔐 Starting login and request capture...")
    api_data, xml_unit_ids, project_name, room_type = login_and_capture_requests()
    if not api_data or not xml_unit_ids:
        print("❌ Failed to capture JSON and XML data.")
        exit(1)
    print("✅ Successfully captured API and XML data.\n")

    print("📜 Loading history from previous runs...")
    history = load_history()
    print("✅ History loaded.\n")

    print("📊 Comparing current availability with previous run...")
    availability_set = set(api_data.get("availabilitySet", []))
    newly_taken, prev_timestamp = update_history(history, availability_set, xml_unit_ids)
    print(f"✅ Found {len(newly_taken)} units newly taken since last run.\n")

    print("📁 Generating Excel workbook with availability data...")
    create_excel(api_data, xml_unit_ids, newly_taken, prev_timestamp, project_name, room_type)
